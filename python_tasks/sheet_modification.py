import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
print(wb)
sheet = wb['Sheet1']

for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value*0.9
    corrected_price_cell = sheet.cell(row,3)
    corrected_price_cell.value  = corrected_price

wb.save('trsactions2.xlsx')
